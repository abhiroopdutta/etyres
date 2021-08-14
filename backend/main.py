import openpyxl
import MySQLdb
import re 
	
f = open("password.txt", "r")
pwd = str(f.read().replace('\n', ''))
f.close()
database = MySQLdb.connect (host="localhost", user = "root", passwd = pwd, db = "tyres")
cursor = database.cursor()

query = """INSERT INTO pv_tyres (ITEM_DESCRIPTION, SIZE, ITEM_CODE, NDP, FRT, SPD, PLSD, TAX_VAL, GST, TOTAL, TCS, INV_VALUE, QUANTITY) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

wb = openpyxl.load_workbook('tyres.xlsx', data_only='True')
tyres_xl = wb['Sheet2']

n_rows = tyres_xl.max_row



for i in range(2, tyres_xl.max_row):
	ITEM_DESCRIPTION = tyres_xl.cell(row=i, column=1).value
	
	split_words = ITEM_DESCRIPTION.split(' ', 2)
	SIZE = re.sub("[^0-9]", "", split_words[0])
	if "R" in split_words[1]:
		SIZE = re.sub("[^0-9]", "", split_words[0]+split_words[1])
	
	ITEM_CODE = tyres_xl.cell(row=i, column=2).value
	NDP = tyres_xl.cell(row=i, column=3).value
	FRT = tyres_xl.cell(row=i, column=4).value
	SPD = tyres_xl.cell(row=i, column=5).value
	PLSD = tyres_xl.cell(row=i, column=6).value
	TAX_VAL = tyres_xl.cell(row=i, column=7).value
	GST = tyres_xl.cell(row=i, column=8).value
	TOTAL = tyres_xl.cell(row=i, column=9).value
	TCS = tyres_xl.cell(row=i, column=10).value
	INV_VALUE = tyres_xl.cell(row=i, column=11).value
	QUANTITY = 0
	
	values = (ITEM_DESCRIPTION, SIZE, ITEM_CODE, NDP, FRT, SPD, PLSD, TAX_VAL, GST, TOTAL, TCS, INV_VALUE, QUANTITY)
	cursor.execute(query, values)
	
cursor.close()
database.commit()
database.close()
	
	
	
	

	

