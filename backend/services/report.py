from mongoengine import Q
import datetime
import openpyxl
import os
import json
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE, FORMAT_NUMBER, FORMAT_NUMBER_00
from openpyxl.cell import WriteOnlyCell
from gsttable import compute_gst_tables
from services.product import compute_size
from models import Product, Sale
from services.sale import sale_service
from services.purchase import purchase_service

def get_gst_state_codes():
	file_path = "/app/services/gstStateCodes.json"
	f = open(file_path, "r")
	gst_state_codes = json.loads(f.read())
	f.close()
	return gst_state_codes

# helper function for export_gstr1_report
def round_to_two(sheet, value):
    cell = WriteOnlyCell(sheet, value=round(value, 2))
    cell.number_format = FORMAT_NUMBER_00
    return cell

def prepare_gstr1_report(start_date, end_date):
    start_datetime = datetime.datetime.strptime(start_date + " " + "00:00:00", '%Y-%m-%d %H:%M:%S')
    end_datetime = datetime.datetime.strptime(end_date + " " + "23:59:59", '%Y-%m-%d %H:%M:%S')

    b2b_data_pipeline = [
        # Stage 1: Filter invoices according to date range, invoiceStatus = "paid", GST customers
        {
           "$match":
            {
                "invoiceDate": { "$gte": start_datetime, "$lte": end_datetime },
                "invoiceStatus": "paid",
                "customerDetails.GSTIN": { "$ne": "" } 
            }
        },
        # Stage 2: Concat products[] to form items, drop unnecessary fields
        { 
            "$project": 
            {   
                "_id": 0,
                "invoiceNumber": 1,
                "invoiceDate": 1,
                "invoiceTotal": 1,
                "customerName": "$customerDetails.name",
                "customerGSTIN": "$customerDetails.GSTIN",
                "items": "$productItems"
            } 
        },
        # Stage 3: Unwind (denormalize) items list
        {
            "$unwind" : "$items" 
        },
        # Stage 4: Compute new fields (taxable value, item rate), drop unnecessary fields from items
        { 
            "$project": 
            {   
                "invoiceNumber": 1,
                "invoiceDate": 1,
                "invoiceTotal": 1,
                "customerName": 1,
                "customerGSTIN": 1,

                #new computed fields from items
                "itemRate": {"$round": [{"$multiply": [{ "$sum": [ "$items.CGST", "$items.SGST", "$items.IGST" ] }, 100.00]}, 2]},
                "taxableValue": {"$round": [{ "$multiply": [ "$items.ratePerItem", "$items.quantity" ] }, 2]}
            } 
        },
        # Stage 5 : Sort by invoiceNumber
        {   
            "$sort": { "invoiceNumber": 1} 
        },
        # Stage 6: Group by invoiceNumber and itemRate, sum(taxableValue)
        {
            "$group": {
            "_id": {
                "invoiceNumber": "$invoiceNumber", 
                "itemRate": "$itemRate"
                },
            "invoiceDate": {"$first": "$invoiceDate"},
            "invoiceTotal": {"$first": "$invoiceTotal"},
            "customerName": {"$first": "$customerName"},
            "customerGSTIN": {"$first": "$customerGSTIN"},
            "taxableValue": { "$sum": "$taxableValue"}
            }
        },
        # Stage 7 : Sort by invoiceDate
        {   
            "$sort": { "invoiceDate": 1} 
        },
        # Stage 8: Project stage for reformatting data for further stages
        {
            "$project": {   
                "_id": 0,
                "invoiceNumber": "$_id.invoiceNumber",
                "invoiceDate": "$invoiceDate",
                "invoiceTotal": "$invoiceTotal",
                "customerName": "$customerName",
                "customerGSTIN": "$customerGSTIN",
                "itemRate": "$_id.itemRate",
                "taxableValue": "$taxableValue"
            } 
        },
        # Stage 9 : Create 2 branches, one for invoices (above data) and one for summary
        {
            "$facet": {
                "b2bInvoices": [{
                    "$project": {   
                        "invoiceNumber": 1,
                        "invoiceDate": 1,
                        "invoiceTotal": 1,
                        "customerName": 1,
                        "customerGSTIN": 1,
                        "itemRate": 1,
                        "taxableValue": 1
                    } 
                }],

                "countUniqB2bCustomer": [
                    {
                        "$group": {
                        "_id": "$customerGSTIN",
                        }
                    },
                    {
                        "$count": "value"
                    }
                ],
                "countUniqB2bInvoices": [
                    {
                        "$group": {
                        "_id": "$invoiceNumber",
                        }
                    },
                    {
                        "$count": "value"
                    }
                ],
                "sumInvoiceTotal": [
                    {
                        "$group": {
                        "_id": {"invoiceNumber": "$invoiceNumber", "invoiceTotal": "$invoiceTotal"},
                        }
                    },
                    {
                        "$group": {
                        "_id": "null",
                        "value": { "$sum": "$_id.invoiceTotal"}
                        }
                    }
                ],
                "sumTaxableValue": [
                    {
                        "$group": {
                        "_id": "null",
                        "value": { "$sum": "$taxableValue"}
                        }
                    }
                ],

            }
        }
    ]

    b2c_data_pipeline = [
        # Stage 1: Filter invoices according to date range, invoiceStatus = "paid", non GST customers
        {
           "$match":
            {
                "invoiceDate": { "$gte": start_datetime, "$lte": end_datetime },
                "invoiceStatus": "paid",
                "customerDetails.GSTIN": { "$eq": "" } 
            }
        },
        # Stage 2: drop unnecessary fields
        { 
            "$project": 
            {   
                "_id": 0,
                "items": "$productItems"
            } 
        },
        # Stage 3: Unwind (denormalize) items list
        {
            "$unwind" : "$items" 
        },
        # Stage 4: Compute new fields (taxable value, item rate), drop unnecessary fields from items
        { 
            "$project": 
            {   
                #new computed fields from items
                "itemRate": {"$round": [{"$multiply": [{ "$sum": [ "$items.CGST", "$items.SGST", "$items.IGST" ] }, 100.00]}, 2]},
                "taxableValue": {"$round": [{ "$multiply": [ "$items.ratePerItem", "$items.quantity" ] }, 2]}
            } 
        },
        # Stage 5: Group by rate, sum (taxableValue)
        {
            "$group": 
            {
                "_id": "$itemRate",
                "taxableValue": { "$sum": "$taxableValue"}
            }
        },
        # Stage 6: Create 2 branches, one for invoices (above data) and one for summary
        {
            "$facet": {
                "b2cInvoices": [{
                    "$project": {
                        "_id": 0,   
                        "rate": "$_id",
                        "taxableValue": 1
                    } 
                }],

                "sumTaxableValue": [
                    {
                        "$group": {
                        "_id": "null",
                        "value": { "$sum": "$taxableValue"}
                        }
                    }
                ],

            }
        }
    ]
    
    hsn_data_pipeline = [
        # Stage 1: Filter invoices according to date range, invoiceStatus = "paid"
        {
           "$match":
            {
                "invoiceDate": { "$gte": start_datetime, "$lte": end_datetime },
                "invoiceStatus": "paid",
            }
        },
        # Stage 2: drop unnecessary fields
        { 
            "$project": 
            {   
                "_id": 0,
                "items": "$productItems"
            } 
        },
        # Stage 3: Unwind (denormalize) items list
        {
            "$unwind" : "$items" 
        },
        # Stage 4: Fill IGST values with 0.00 for service items that don't have IGST field 
        {
            "$fill":
            {
            "output":
                {
                    "items.IGST": { "value": 0.00 }
                }
            }
        },
        # Stage 5: Compute new fields (taxable value, item rate), drop unnecessary fields from items
        { 
            "$project": 
            {   
                #new computed fields from items
                "hsn": "$items.HSN",
                "rate": {"$round": [{"$multiply": [{ "$sum": [ "$items.CGST", "$items.SGST", "$items.IGST" ] }, 100.00]}, 2]},
                "quantity": "$items.quantity",
                "taxableValue": {"$round": [{ "$multiply": [ "$items.ratePerItem", "$items.quantity" ] }, 2]},
                "CGST": "$items.CGST",
                "SGST": "$items.SGST",
                "IGST": "$items.IGST",
            } 
        },
        # Stage 6: Compute new fields (CGST, SGST, IGST Amount)
        { 
            "$project": 
            {   
                #new computed fields from items
                "hsn": 1,
                "rate": 1,
                "quantity": 1,
                "taxableValue": 1,
                "CGSTAmount": {"$round": [{"$multiply": ["$CGST", "$taxableValue"]}, 2]},
                "SGSTAmount": {"$round": [{"$multiply": ["$SGST", "$taxableValue"]}, 2]},
                "IGSTAmount": {"$round": [{"$multiply": ["$IGST", "$taxableValue"]}, 2]},
            } 
        },
        # Stage 7: Group by hsn, sum (CGST), sum(SGST), sum(IGST)
        {
            "$group": 
            {
                "_id": {"hsn": "$hsn", "rate": "$rate"},
                "quantity": { "$sum": "$quantity"},
                "taxableValue": { "$sum": "$taxableValue"},
                "CGSTAmount": { "$sum": "$CGSTAmount"},
                "SGSTAmount": { "$sum": "$SGSTAmount"},
                "IGSTAmount": { "$sum": "$IGSTAmount"},
            }
        },
        # Stage 8: Create 2 branches, one for invoices (above data) and one for summary
        {
            "$facet": {
                "hsnItems": [{
                    "$project": {
                        "_id": 0,   
                        "hsn": "$_id.hsn",
                        "rate": "$_id.rate",
                        "quantity": 1,
                        "taxableValue": 1,
                        "CGSTAmount": 1,
                        "SGSTAmount": 1,
                        "IGSTAmount": 1,
                    } 
                }],

                "countUniqHSN": [
                    {
                        "$count": "value"
                    }
                ],
                "sumTaxableValue": [
                    {
                        "$group": {
                        "_id": "null",
                        "value": { "$sum": "$taxableValue"}
                        }
                    }
                ],
                "sumCGSTAmount": [
                    {
                        "$group": {
                        "_id": "null",
                        "value": { "$sum": "$CGSTAmount"}
                        }
                    }
                ],
                "sumSGSTAmount": [
                    {
                        "$group": {
                        "_id": "null",
                        "value": { "$sum": "$SGSTAmount"}
                        }
                    }
                ],
                "sumIGSTAmount": [
                    {
                        "$group": {
                        "_id": "null",
                        "value": { "$sum": "$IGSTAmount"}
                        }
                    }
                ]

            }
        }
    
    ]

    docs_data_pipeline = [
        {
            "$match": {
                "invoiceDate": { "$gte": start_datetime, "$lte": end_datetime },
            }
        },
        {
            "$facet": {
                "minInvoiceNumber": [
                    {
                        "$group": {
                            "_id": "null",
                            "value": { "$min": "$invoiceNumber" },
                        }
                    },
                ],
                "maxInvoiceNumber": [
                    {
                        "$group": {
                            "_id": "null",
                            "value": { "$max": "$invoiceNumber" }
                        }
                    },
                ],
                "paidInvoicesCount": [
                    {
                        "$match":{"invoiceStatus": "paid",}
                    },
                    {
                        "$count": "value"
                    }
                ],
                "cancelledInvoicesCount": [
                    {
                        "$match":{"invoiceStatus": "cancelled",}
                    },
                    {
                        "$count": "value"
                    }
                ]
            }
        }
    ]

    b2b_data = list(Sale.objects().aggregate(b2b_data_pipeline))[0]
    b2c_data = list(Sale.objects().aggregate(b2c_data_pipeline))[0]
    hsn_data = list(Sale.objects().aggregate(hsn_data_pipeline))[0]
    docs_data = list(Sale.objects().aggregate(docs_data_pipeline))[0]
    return {
        "b2bData": {
            "b2bSummary": {
                "countUniqB2bCustomer": b2b_data["countUniqB2bCustomer"][0]["value"],
                "countUniqB2bInvoices": b2b_data["countUniqB2bInvoices"][0]["value"],
                "sumInvoiceTotal": b2b_data["sumInvoiceTotal"][0]["value"],
                "sumTaxableValue": b2b_data["sumTaxableValue"][0]["value"]
            },
            "b2bItems": b2b_data["b2bInvoices"]
        },
        "b2cData": {
            "b2cSummary": {
                "sumTaxableValue": b2c_data["sumTaxableValue"][0]["value"]
            },
            "b2cItems": b2c_data["b2cInvoices"]
        },
        "hsnData": {
            "hsnSummary": {
                "countUniqHSN": hsn_data["countUniqHSN"][0]["value"],
                "sumTaxableValue": hsn_data["sumTaxableValue"][0]["value"],
                "sumCGSTAmount": hsn_data["sumCGSTAmount"][0]["value"],
                "sumSGSTAmount": hsn_data["sumSGSTAmount"][0]["value"],
                "sumIGSTAmount": hsn_data["sumIGSTAmount"][0]["value"],
            },
            "hsnItems": hsn_data["hsnItems"]
        },
        "docsData": {
            "docsSummary": {
                "totalDocs": docs_data["paidInvoicesCount"][0]["value"],
                "totalcancelledDocs": docs_data["cancelledInvoicesCount"][0]["value"] if docs_data["cancelledInvoicesCount"] else 0

            },
            "docsItems": [{
                "minInvoiceNumber": docs_data["minInvoiceNumber"][0]["value"],
                "maxInvoiceNumber": docs_data["maxInvoiceNumber"][0]["value"],
                "paidInvoicesCount": docs_data["paidInvoicesCount"][0]["value"],
                "cancelledInvoicesCount": docs_data["cancelledInvoicesCount"][0]["value"] if docs_data["cancelledInvoicesCount"] else 0
            }],
        },   
    }

def export_gstr1_report(reports_dir, start_date, end_date):
    file_base_dir = reports_dir
    filename = str(datetime.datetime.now())+"gstr1.xlsx"
    filename = ''.join(filter(str.isalnum, filename))
    wb = openpyxl.Workbook(write_only=True)
    gst_state_codes = get_gst_state_codes()
    gstr1_report = prepare_gstr1_report(start_date, end_date)
    # ----------------------------------------------b2b sheet----------------------------------------------
    b2b_sheet = wb.create_sheet("b2b,sez,de")
    b2b_summary = gstr1_report["b2bData"]["b2bSummary"]
    b2b_items = gstr1_report["b2bData"]["b2bItems"]

    excel_b2b_summary = {
            "No. of Recipients" : b2b_summary["countUniqB2bCustomer"],
            "No. of Invoices": b2b_summary["countUniqB2bInvoices"],
            "Total Invoice Value": round_to_two(b2b_sheet, b2b_summary["sumInvoiceTotal"]),
            "Total Taxable Value": round_to_two(b2b_sheet, b2b_summary["sumTaxableValue"]),
            "Total Cess": round_to_two(b2b_sheet, 0.00)
        }
    excel_b2b_items = []
    for b2b_item in b2b_items:
        excel_b2b_items.append({
            "GSTIN/UIN of Recipient": b2b_item["customerGSTIN"],
            "Receiver Name": b2b_item["customerName"],
            "Invoice Number": b2b_item["invoiceNumber"],
            "Invoice date": b2b_item["invoiceDate"].strftime("%d-%b-%Y"), 
            "Invoice Value": round_to_two(b2b_sheet, b2b_item["invoiceTotal"]), 
            "Place Of Supply": f"09-{gst_state_codes['09']}", 
            "Reverse Charge": "N", 
            "Applicable % of Tax Rate": "", 
            "Invoice Type": "Regular B2B", 
            "E-Commerce GSTIN": "", 
            "Rate": round_to_two(b2b_sheet, b2b_item["itemRate"]), 
            "Taxable Value": round_to_two(b2b_sheet, b2b_item["taxableValue"]),
            "Cess Amount": round_to_two(b2b_sheet, 0.00)
        })

    title_row = ["Summary For B2B, SEZ, DE (4A, 4B, 6B, 6C)"] + ["" for i in range(11)] + ["HELP"]
    summary_headers = ( ["No. of Recipients", "", "No. of Invoices", "", "Total Invoice Value"] 
                        + ["" for i in range(6)] 
                        + ["Total Taxable Value", "Total Cess"])
    summary_data = ([excel_b2b_summary["No. of Recipients"], "", excel_b2b_summary["No. of Invoices"], "", 
                     excel_b2b_summary["Total Invoice Value"] ]
                    +["" for i in range(6)]
                    +[excel_b2b_summary["Total Taxable Value"], excel_b2b_summary["Total Cess"]])
    column_headers = list(excel_b2b_items[0])

    b2b_sheet.append(title_row)
    b2b_sheet.append(summary_headers)
    b2b_sheet.append(summary_data)
    b2b_sheet.append(column_headers)
    for excel_b2b_item in excel_b2b_items:
        b2b_sheet.append(list(excel_b2b_item.values()))

    # ----------------------------------------------b2cs sheet----------------------------------------------
    b2c_sheet = wb.create_sheet("b2cs")
    b2c_summary = gstr1_report["b2cData"]["b2cSummary"]
    b2c_items = gstr1_report["b2cData"]["b2cItems"]
    
    excel_b2c_summary = {
            "Total Taxable  Value" : round_to_two(b2c_sheet, b2c_summary["sumTaxableValue"]),
            "Total Cess": round_to_two(b2c_sheet, 0.00)
        }
    excel_b2c_items = []
    for b2c_item in b2c_items:
        excel_b2c_items.append({
            "Type": "OE",
            "Place Of Supply": f"09-{gst_state_codes['09']}", 
            "Applicable % of Tax Rate": "", 
            "Rate": round_to_two(b2c_sheet, b2c_item["rate"]), 
            "Taxable Value": round_to_two(b2c_sheet, b2c_item["taxableValue"]),
            "Cess Amount": round_to_two(b2c_sheet, 0.00),
            "E-Commerce GSTIN": ""
        })

    title_row = ["Summary For B2CS(7)"] + ["" for i in range(5)] + ["HELP"]

    summary_headers = ["" for i in range(4)] + ["Total Taxable  Value", "Total Cess", ""]
    summary_data = ["" for i in range(4)] + [excel_b2c_summary["Total Taxable  Value"], excel_b2c_summary["Total Cess"], ""]
    column_headers = list(excel_b2c_items[0])

    b2c_sheet.append(title_row)
    b2c_sheet.append(summary_headers)
    b2c_sheet.append(summary_data)
    b2c_sheet.append(column_headers)
    for excel_b2c_item in excel_b2c_items:
        b2c_sheet.append(list(excel_b2c_item.values()))

# ----------------------------------------------hsn sheet----------------------------------------------
    hsn_sheet = wb.create_sheet("hsn")
    hsn_summary = gstr1_report["hsnData"]["hsnSummary"]
    hsn_items = gstr1_report["hsnData"]["hsnItems"]
    
    excel_hsn_summary = {
            "No. of HSN" : hsn_summary["countUniqHSN"],
            "Total Value": "",
            "Total Taxable Value": round_to_two(hsn_sheet, hsn_summary["sumTaxableValue"]),
            "Total Integrated Tax": round_to_two(hsn_sheet, hsn_summary["sumIGSTAmount"]),
            "Total Central Tax": round_to_two(hsn_sheet, hsn_summary["sumCGSTAmount"]),
            "Total State/UT Tax": round_to_two(hsn_sheet, hsn_summary["sumSGSTAmount"]),
            "Total Cess": round_to_two(hsn_sheet, 0.00)
        }
    excel_hsn_items = []
    for item in hsn_items:
        excel_hsn_items.append({
            "HSN": item["hsn"],
            "Description": "", 
            "UQC": "NOS-NUMBERS", 
            "Total Quantity": round_to_two(hsn_sheet, item["quantity"]), 
            "Total Value": "",
            "Rate": round_to_two(hsn_sheet, item["rate"]),
            "Taxable Value": round_to_two(hsn_sheet, item["taxableValue"]),
            "Integrated Tax Amount": round_to_two(hsn_sheet, item["IGSTAmount"]),
            "Central Tax Amount" : round_to_two(hsn_sheet, item["CGSTAmount"]),
            "State/UT Tax Amount": round_to_two(hsn_sheet, item["SGSTAmount"]),
            "Cess Amount": round_to_two(hsn_sheet, 0.00)
        })

    title_row = ["Summary For HSN(12)"] + ["" for i in range(9)] + ["HELP"]
    summary_headers = [
        "No. of HSN", 
        "", 
        "", 
        "", 
        "Total Value", 
        "", 
        "Total Taxable Value", 
        "Total Integrated Tax",
        "Total Central Tax",
        "Total State/UT Tax",
        "Total Cess"
        ]
    summary_data = [
        excel_hsn_summary["No. of HSN"], 
        "", 
        "", 
        "", 
        excel_hsn_summary["Total Value"], 
        "", 
        excel_hsn_summary["Total Taxable Value"], 
        excel_hsn_summary["Total Integrated Tax"],
        excel_hsn_summary["Total Central Tax"],
        excel_hsn_summary["Total State/UT Tax"],
        excel_hsn_summary["Total Cess"]
        ]
    column_headers = list(excel_hsn_items[0])

    hsn_sheet.append(title_row)
    hsn_sheet.append(summary_headers)
    hsn_sheet.append(summary_data)
    hsn_sheet.append(column_headers)
    for excel_hsn_item in excel_hsn_items:
        hsn_sheet.append(list(excel_hsn_item.values()))

# ----------------------------------------------docs sheet----------------------------------------------
    docs_sheet = wb.create_sheet("docs")
    docs_summary = gstr1_report["docsData"]["docsSummary"]
    docs_items = gstr1_report["docsData"]["docsItems"]
    
    excel_docs_summary = {
            "Total Number" : docs_summary["totalDocs"],
            "Total Cancelled": docs_summary["totalcancelledDocs"],
        }
    excel_docs_items = []
    for item in docs_items:
        excel_docs_items.append({
            "Nature of Document": "Invoices for outward supply",
            "Sr. No. From": item["minInvoiceNumber"], 
            "Sr. No. To": item["maxInvoiceNumber"], 
            "Total Number": item["paidInvoicesCount"], 
            "Cancelled": item["cancelledInvoicesCount"],
        })

    title_row = ["Summary of documents issued during the tax period (13)"] + ["" for i in range(3)] + ["HELP"]
    summary_headers = [
        "", 
        "", 
        "", 
        "Total Number", 
        "Total Cancelled", 
        ]
    summary_data = [
        "", 
        "", 
        "", 
        excel_docs_summary["Total Number"], 
        excel_docs_summary["Total Cancelled"], 
        ]
    column_headers = list(excel_docs_items[0])

    docs_sheet.append(title_row)
    docs_sheet.append(summary_headers)
    docs_sheet.append(summary_data)
    docs_sheet.append(column_headers)
    for excel_docs_item in excel_docs_items:
        docs_sheet.append(list(excel_docs_item.values()))

    wb.save(file_base_dir+filename)
    return filename

def export_purchase_report(reports_dir, invoices):
    file_base_dir = reports_dir
    filename = str(datetime.datetime.now())+"purchase_report.xlsx"
    wb = openpyxl.Workbook() 
    sheet = wb.active
    sheet.title = "Purchase Report"
    sheet.freeze_panes = 'A2'

    column_headers = [
        "Invoice No.", 
        "Invoice Date", 
        "Invoice Status", 
        "Claim Invoice", 
        "Supplier Name", 
        "Supplier GSTIN", 
        "Item Description", 
        "Item Code", 
        "HSN", 
        "Qty", 
        "Rate per Item", 
        "Taxable Val", 
        "CGST Rate", 
        "CGST Amt", 
        "SGST Rate", 
        "SGST Amt", 
        # "IGST Rate", 
        # "IGST Amt", 
        "Total",
        "size"]

    item_keys = [
        "itemDesc",
        "itemCode",
        "HSN",
        "quantity",
        "ratePerItem",
        "taxableValue",
        "CGST",
        "CGSTAmount",
        "SGST",
        "SGSTAmount",
        # "IGST",
        # "IGSTAmount",
        "value",
    ]
    
    for i, column_header in enumerate(column_headers):
        sheet.cell(row=1, column=i+1).value = column_header

    #find a better way to do the following, instead of using base index, use sheet.max_row ?
    base_index = 0
    for invoice in invoices:
        gst_tables = compute_gst_tables(invoice.items)
        tax_table = gst_tables["GST_table"]
        products = tax_table["products"]

        total_items = len(products)
        for i, product in enumerate(products):
            row_index = i+2+base_index
            sheet.cell(row = row_index, column = 1).value = invoice.invoiceNumber
            sheet.cell(row = row_index, column = 2).value = invoice.invoiceDate.strftime("%d/%m/%Y")
            sheet.cell(row = row_index, column = 3).value = invoice.invoiceStatus
            sheet.cell(row = row_index, column = 4).value = "No" if invoice.claimInvoice ==0 else "Yes"
            sheet.cell(row = row_index, column = 5).value = invoice.supplier.name
            sheet.cell(row = row_index, column = 6).value = invoice.supplier.GSTIN

            for j, item_key in enumerate(item_keys):
                sheet.cell(row = row_index, column = 7+j).value = product[item_key]
            
            sheet.cell(row = row_index, column = 7+len(item_keys)).value = compute_size(product["itemDesc"])

        base_index += total_items

    total_row_index = sheet.max_row
    
    # total row label
    sheet.cell(row=total_row_index+1, column=2).value = "TOTAL"
    # total quantity
    total_quantity = '= SUM(J2:J'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=10).value = total_quantity
    # total taxable val
    total_taxable_val = '= SUM(L2:L'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=12).value = total_taxable_val
    # total CGST
    total_CGST = '= SUM(N2:N'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=14).value = total_CGST
    # total SGST
    total_SGST = '= SUM(P2:P'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=16).value = total_SGST
    # total 
    total = '= SUM(Q2:Q'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=17).value = total

    for i in range(2, sheet.max_row+1):
        sheet.cell(row=i, column=6).number_format = FORMAT_NUMBER

    #this adjusts the font and column width of the whole sheet
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), (len(str(cell.value))+1)*1.2))    
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value
    ####

    # Apply styles to header and footer row
    for col in range(1, sheet.max_column + 1):
        cell_header = sheet.cell(1, col)
        cell_header.fill = PatternFill("solid", fgColor="C5C5C5")
        cell_header.border = Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        cell_header.font = Font(name="Calibri", bold=True)
        cell_footer = sheet.cell(sheet.max_row, col)
        cell_footer.fill = PatternFill("solid", fgColor="C5C5C5")
        cell_footer.border = Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        cell_footer.font = Font(name="Calibri", bold=True)

    wb.save(file_base_dir+filename)
    return filename

def export_sales_report(reports_dir, invoices):
    file_base_dir = reports_dir
    filename = str(datetime.datetime.now())+"sales_report.xlsx"
    filename = ''.join(filter(str.isalnum, filename))
    wb = openpyxl.Workbook() 
    sheet = wb.active
    sheet.title = "Sales Report"
    sheet.freeze_panes = 'A2'

    column_headers = [
        "Invoice No.", 
        "Invoice Date", 
        "Customer Name", 
        "GSTIN", 
        "Item Description", 
        "Item Code", 
        "HSN", 
        "Qty", 
        "Rate per Item", 
        "Taxable Val", 
        "CGST Rate", 
        "CGST Amt", 
        "SGST Rate", 
        "SGST Amt", 
        "IGST Rate", 
        "IGST Amt", 
        "Total",
        "size"]

    item_keys = [
        "itemDesc",
        "itemCode",
        "HSN",
        "quantity",
        "ratePerItem",
        "taxableValue",
        "CGST",
        "CGSTAmount",
        "SGST",
        "SGSTAmount",
        "IGST",
        "IGSTAmount",
        "value",
    ]

    for i, column_header in enumerate(column_headers):
        sheet.cell(row=1, column=i+1).value = column_header

    #find a better way to do the following, instead of using base index, use sheet.max_row ?
    base_index = 0
    for invoice in invoices:
        gst_tables = compute_gst_tables(invoice.productItems)
        tax_table = gst_tables["GST_table"]
        # if (invoice.customerDetails.POS.startswith("09")):
        #     tax_table = gst_tables["GST_table"]
        # else:
        #     tax_table = gst_tables["IGST_table"]
        products = tax_table["products"]

        total_items = len(products)
        for i, product in enumerate(products):
            row_index = i+2+base_index
            sheet.cell(row = row_index, column = 1).value = invoice.invoiceNumber
            sheet.cell(row = row_index, column = 2).value = invoice.invoiceDate.strftime("%d/%m/%Y")
            sheet.cell(row = row_index, column = 3).value = invoice.customerDetails.name
            sheet.cell(row = row_index, column = 4).value = invoice.customerDetails.GSTIN

            for j, item_key in enumerate(item_keys):
                sheet.cell(row = row_index, column = 5+j).value = product[item_key]
            
            sheet.cell(row = row_index, column = 5+len(item_keys)).value = compute_size(product["itemDesc"])
            
        base_index += total_items

    total_row_index = sheet.max_row
    
    # total row label
    sheet.cell(row=total_row_index+1, column=7).value = "TOTAL"
    # total quantity
    total_quantity = '= SUM(H2:H'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=8).value = total_quantity
    # total taxable val
    total_taxable_val = '= SUM(J2:J'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=10).value = total_taxable_val
    # total CGST
    total_CGST = '= SUM(L2:L'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=12).value = total_CGST
    # total SGST
    total_SGST = '= SUM(N2:N'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=14).value = total_SGST
    # total IGST
    total_IGST = '= SUM(P2:P'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=16).value = total_IGST
    # total 
    total = '= SUM(Q2:Q'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=17).value = total

    for i in range(2, sheet.max_row+1):
        sheet.cell(row=i, column=11).number_format = FORMAT_PERCENTAGE
        sheet.cell(row=i, column=13).number_format = FORMAT_PERCENTAGE
        sheet.cell(row=i, column=15).number_format = FORMAT_PERCENTAGE

    #this adjusts the font and column width of the whole sheet
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), (len(str(cell.value)))*1.2))    
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value
    ####

    # Apply styles to header and footer row
    for col in range(1, sheet.max_column + 1):
        cell_header = sheet.cell(1, col)
        cell_header.fill = PatternFill("solid", fgColor="b4f05c")
        cell_header.border = Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        cell_header.font = Font(name="Calibri", bold=True)
        cell_footer = sheet.cell(sheet.max_row, col)
        cell_footer.fill = PatternFill("solid", fgColor="b4f05c")
        cell_footer.border = Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        cell_footer.font = Font(name="Calibri", bold=True)

    wb.save(file_base_dir+filename)
    return filename

def stock_report(reports_dir):
    products = Product.objects()
    file_base_dir = reports_dir
    filename = (str(datetime.datetime.now())+"stock_report.xlsx")
    filename = ''.join(filter(str.isalnum, filename))
    wb = openpyxl.Workbook() 
    sheet = wb.active

    sheet.freeze_panes = 'A2'
    column_headers = ["Item Description", "Item Code", "HSN", "GST", "category", "Size", "Cost Price", "Stock"]
    product_keys = [
        "itemDesc",
        "itemCode",
        "HSN",
        "GST",
        "category",
        "size",
        "costPrice",
        "stock"
    ]

    for i, column_header in enumerate(column_headers):
        sheet.cell(row=1, column=i+1).value = column_header

    for i, product in enumerate(products):
        for j, product_key in enumerate(product_keys):
            sheet.cell(row = i+2, column = j+1).value = product[product_key]

    #this adjusts the font and column width of the whole sheet
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), (len(str(cell.value))+0)*1.2))    
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value
    ####

    # Apply styles to header and footer row
    for col in range(1, sheet.max_column + 1):
        cell_header = sheet.cell(1, col)
        cell_header.fill = PatternFill("solid", fgColor="fcfa8e")
        cell_header.border = Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        cell_header.font = Font(name="Calibri", bold=True)

    wb.save(file_base_dir+filename)
    return filename

class ReportService:
    def create_report(self, reports_dir, report_req_info):
        os.makedirs(reports_dir, exist_ok = True) #make the dir if it doesn't exist
        if report_req_info["reportType"] == "stock":
            return stock_report(reports_dir)
        elif report_req_info["reportType"] == "sale":
            results = sale_service.get_invoices(**report_req_info["query"])
            if report_req_info["exportType"] == "regular":
                return export_sales_report(reports_dir, results["data"])
            elif report_req_info["exportType"] == "gstr1":
                return export_gstr1_report(reports_dir, report_req_info["query"]["start"], report_req_info["query"]["end"])
                    
        elif report_req_info["reportType"] == "purchase":
            results = purchase_service.get_invoices(**report_req_info["query"])
            return export_purchase_report(reports_dir, results["data"])

report_service = ReportService()