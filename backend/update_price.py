import openpyxl
import re, os
from models import Product
import json

default_pv_price_details = {
	"passenger car":{"tyre_freight":20, "tube_freight":3, "spd":0.015, "plsd":0.025}, 
	"2 wheeler":{"tyre_freight":6, "tube_freight":3, "spd":0.015, "plsd":0.025}, 
	"3 wheeler":{"tyre_freight":8, "tube_freight":3, "spd":0.014, "plsd":0.025},
	"scv":{"tyre_freight":40, "tube_freight":5, "spd":0.014, "plsd":0.025},
	"tubeless valve":{ "valve_freight":0, "spd":0.0, "plsd":0.0}
	}

hsn_gst = {
	"tyre": {
		"HSN": "4011",
		"GST": 0.28
	},
	"tube": {
		"HSN": "4013",
		"GST": 0.28
	},
	"flap": {
		"HSN": "4012",
		"GST": 0.28
	},
	"valve": {
		"HSN": "8481",
		"GST": 0.18
	}
}

def get_pv_price_details():
	file_path = "./pv_price_details.json"

	# if file doesn't exist, create it using default values
	if not os.path.isfile(file_path):
		with open(file_path, 'w',  encoding='utf-8') as outfile:
			json.dump(default_pv_price_details, outfile, ensure_ascii=False, indent=4)	

	f = open(file_path, "r")
	pv_price_details = json.loads(f.read())
	f.close()

	pv_price_details_list = []
	for key, value in pv_price_details.items():
		if key == "tubeless valve":
			continue
		item = {}
		item["vehicleType"] = key
		item["spd"] = value["spd"]
		item["plsd"] = value["plsd"]
		item["tyreFreight"] = value["tyre_freight"]
		item["tubeFreight"] = value["tube_freight"]
		pv_price_details_list.append(item)
	return pv_price_details_list

#returns the product type (tyre/tube/etc) depending on item_code
def get_product_type(item_code):
	if (item_code[1] in ["U", "W", "Y"]):
		return "tube"
	elif (item_code == "RR100TR414A"):
		return "valve"
	else:
		return "tyre"

#calculates cost_price based on item_code, net_ndp, vehicle_type
def compute_price(item_code, net_ndp, freight, spd_rate, plsd_rate):
	spd = float(spd_rate*net_ndp)
	plsd = float(net_ndp+freight-spd)*float(plsd_rate)
	taxable_val = float(net_ndp+freight-spd-plsd)
	gst = taxable_val*float(get_gst_rate(item_code))
	cost_price = round(float(taxable_val+gst),2)

	print("\n net ndp:", net_ndp, "freight:", freight, "spd:", spd, "plsd:", plsd, "total discount:", spd+plsd,"taxable val:", taxable_val, "gst:", gst, "cost price:", cost_price)
	return cost_price

def compute_size(item_desc):
	if (item_desc == "TR 414 TUBELES TYRE VALVE -D"):
		return "valve"
	else:
		digits = re.findall('\d+', item_desc)
		return "".join(digits)

#returns HSN Code depending on item_code
def get_hsn(item_code):
	product_type = get_product_type(item_code)
	return hsn_gst[product_type]["HSN"]

#returns GST rate depending on item_code
def get_gst_rate(item_code):
	product_type = get_product_type(item_code)
	gst_rate = float(hsn_gst[product_type]["GST"])
	return gst_rate

#categorizes items into segment+tyre or segment+tube
def categorize(vehicle_type, item_code):
	product_type = get_product_type(item_code)
	if (product_type == "tube"):
		return vehicle_type.replace(" ", "_")+"_tube"
	elif (product_type == "valve"):
		return vehicle_type.replace(" ", "_")
	else:
		return vehicle_type.replace(" ", "_")+"_tyre"

#check if item exists, then update the price, else insert the item as new item by computing all columns
def load_to_db(vehicle_type, item_desc, item_code, cost_price):
	#if item already exists then update the price only
	if (Product.objects(itemCode=item_code).first()):
		Product.objects(itemCode=item_code).first().update(costPrice=cost_price)
	#else compute all fields and add the item
	else:
		size = compute_size(item_desc)
		hsn = get_hsn(item_code)
		category = categorize(vehicle_type, item_code)
		gst = get_gst_rate(item_code)
		Product(itemDesc = item_desc, 
				itemCode = item_code, 
				HSN = hsn, 
				GST = gst, 
				category = category, 
				size = size, 
				costPrice = cost_price, 
				stock = 0).save()
	
	return 0

# what if given file is not excel, to do: add validation of file
def update_price(pv_price_details_list, file):
	pv_price_details = {}
	for item in pv_price_details_list:
		new_item = {}
		new_item["spd"] = float(item["spd"])
		new_item["plsd"] = float(item["plsd"])
		new_item["tyre_freight"] = float(item["tyreFreight"])
		new_item["tube_freight"] = float(item["tubeFreight"])
		pv_price_details[item["vehicleType"]] = new_item
	pv_price_details["tubeless valve"] = {"spd":0, "plsd":0, "valve_freight":0}

	other_price_details = {
	"truck and bus":{"tyre_freight":0, "tube_freight":0, "spd":0.0, "plsd":0.0}, 
	"farm":{"tyre_freight":0, "tube_freight":0, "spd":0.0, "plsd":0.0}, 
	"lcv":{"tyre_freight":0, "tube_freight":0, "spd":0.0, "plsd":0.0}, 
	"tt":{"tyre_freight":0, "tube_freight":0, "spd":0.0, "plsd":0.0}, 
	"industrial":{"tyre_freight":0, "tube_freight":0, "spd":0.0, "plsd":0.0}, 
	"earthmover":{"tyre_freight":0, "tube_freight":0, "spd":0.0, "plsd":0.0}, 
	"jeep":{"tyre_freight":0, "tube_freight":0, "spd":0.0, "plsd":0.0}, 
	"loose tube/flaps":{"tyre_freight":0, "tube_freight":0, "spd":0.0, "plsd":0.0},
	"adv":{"tyre_freight":0, "tube_freight":0, "spd":0.0, "plsd":0.0}
	}

	# save pv price details to a file
	with open('pv_price_details.json', 'w',  encoding='utf-8') as outfile:
		json.dump(pv_price_details, outfile, ensure_ascii=False, indent=4)

	wb = openpyxl.load_workbook(file, data_only='True')
	tyres_xl = wb['Sheet1']	
	vehicle_type = ""
	for i in range(1,tyres_xl.max_row+1):
		column_a = tyres_xl.cell(row=i, column=1).value
		
		# skip the row if column_a is empty
		if (not column_a):
			continue

		column_a = str(column_a)
		print("\nrow", i, "column1", column_a)

		# if this row is a vehicle type row, skip it after getting the vehicle type
		if (column_a.lower().strip() in pv_price_details) or (column_a.lower().strip() in other_price_details):
			vehicle_type = column_a.lower().strip()
			continue

		# skip this row if its not a vehicle type row, neither it contains product info
		if (not vehicle_type):
			continue

		item_desc = str(tyres_xl.cell(row=i, column=1).value).strip()
		item_code = str(tyres_xl.cell(row=i, column=2).value).strip()[:-2]
		net_ndp = round(float(str(tyres_xl.cell(row=i, column=5).value)))

		if (vehicle_type in pv_price_details):
			spd_rate = pv_price_details[vehicle_type]["spd"]
			plsd_rate = pv_price_details[vehicle_type]["plsd"]
			product_type = get_product_type(item_code)
			if (product_type == "tube"):
				freight = pv_price_details[vehicle_type]["tube_freight"]
			elif (product_type == "valve"):
				freight = pv_price_details[vehicle_type]["valve_freight"]
			else:
				freight = pv_price_details[vehicle_type]["tyre_freight"]
			cost_price = compute_price(item_code, net_ndp, freight, spd_rate, plsd_rate)

		elif (vehicle_type in other_price_details):
			cost_price = 0.0

		load_to_db(vehicle_type, item_desc, item_code, cost_price)

	os.remove(file)