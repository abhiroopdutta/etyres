from create_order import compute_gst_tables
from update_price import compute_size
from mongo_agg_pipelines import prepare_gstr1_report
from models import Purchase, Sale, Product
from mongoengine import Q
import datetime
import openpyxl
import os
import json
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE, FORMAT_NUMBER, FORMAT_NUMBER_00
from openpyxl.cell import WriteOnlyCell
import pandas as pd

def get_sales_report(
    invoiceNumber= "",
    invoiceDateFrom= "",
    invoiceDateTo= "",
    invoiceStatus= ["due", "paid", "cancelled"],
    customerName= "",
    customerContact= "",
    customerVehicleNumber= "",
    customerGSTIN= "",
    pageRequest= 1,
    maxItemsPerPage= 5,
):
    pageRequest = int(pageRequest)
    maxItemsPerPage = int(maxItemsPerPage)
    results = {
        "invoices": [],
        "pagination": { 
            "pageNumber": 0, 
            "pageSize": 0, 
            "totalResults" : 0 }
        
    }
    page_start = (pageRequest-1)*maxItemsPerPage
    page_end = pageRequest*maxItemsPerPage
    query = Q(invoiceNumber__gte=0) # dummy query
    
    if (invoiceNumber.isnumeric()):
        query &= Q(invoiceNumber=int(invoiceNumber))
    if (invoiceStatus):
        query &= Q(invoiceStatus__in=invoiceStatus)
    if (customerName):
        query &= Q(customerDetails__name__icontains=customerName)
    if (customerContact):
        query &= Q(customerDetails__contact__icontains=customerContact)
    if (customerVehicleNumber):
        query &= Q(customerDetails__vehicleNumber__icontains=customerVehicleNumber)
    if (customerGSTIN):
        query &= Q(customerDetails__GSTIN__icontains=customerGSTIN)      
    if (invoiceDateFrom and invoiceDateTo):
        start_datetime = datetime.datetime.strptime(invoiceDateFrom[:10] + " " + "00:00:00", '%Y-%m-%d %H:%M:%S')
        end_datetime = datetime.datetime.strptime(invoiceDateTo[:10] + " " + "23:59:59", '%Y-%m-%d %H:%M:%S')
        query &= Q(invoiceDate__gte=start_datetime) & Q(invoiceDate__lte=end_datetime)

    results["invoices"] = Sale.objects(query).order_by('-_id')[page_start:page_end]
    results["pagination"]["totalResults"] = Sale.objects(query).order_by('-_id')[page_start:page_end].count()
    results["pagination"]["pageNumber"] = pageRequest
    results["pagination"]["pageSize"] = len(results["invoices"])
    return results

def get_purchase_report(
    invoiceNumber= "",
    invoiceDateFrom= "",
    invoiceDateTo= "",
    claimInvoice= "",
    pageRequest= 1,
    maxItemsPerPage= 5,
):
    pageRequest = int(pageRequest)
    maxItemsPerPage = int(maxItemsPerPage)
    results = {
        "invoices": [],
        "pagination": { 
            "pageNumber": 0, 
            "pageSize": 0, 
            "totalResults" : 0 }
        
    }
    page_start = (pageRequest-1)*maxItemsPerPage
    page_end = pageRequest*maxItemsPerPage
    query = Q(invoiceNumber__gte=0) # dummy query
    
    if (invoiceNumber.isnumeric()):
        query &= Q(invoiceNumber=int(invoiceNumber))
    if (claimInvoice):
        if (claimInvoice == "true"):
            query &= Q(claimInvoice=True)
        else:
            query &= Q(claimInvoice=False)
    if (invoiceDateFrom and invoiceDateTo):
        start_datetime = datetime.datetime.strptime(invoiceDateFrom[:10] + " " + "00:00:00", '%Y-%m-%d %H:%M:%S')
        end_datetime = datetime.datetime.strptime(invoiceDateTo[:10] + " " + "23:59:59", '%Y-%m-%d %H:%M:%S')
        query &= Q(invoiceDate__gte=start_datetime) & Q(invoiceDate__lte=end_datetime)

    results["invoices"] = Purchase.objects(query).order_by('-invoiceDate')[page_start:page_end]
    results["pagination"]["totalResults"] = Purchase.objects(query).order_by('-invoiceDate')[page_start:page_end].count()
    results["pagination"]["pageNumber"] = pageRequest
    results["pagination"]["pageSize"] = len(results["invoices"])
    return results

def report_handler(report_req_info):
    os.makedirs("./tempdata/sales_report/", exist_ok = True) #make the dir if it doesn't exist
    if report_req_info["reportType"] == "stock":
        return stock_report()
    elif report_req_info["reportType"] == "sale":
        results = get_sales_report(report_req_info["filters"], 
                            report_req_info["sorters"], 
                            report_req_info["pageRequest"], 
                            report_req_info["maxItemsPerPage"])
        if (report_req_info["export"]["required"]):
            if report_req_info["export"]["type"] == "regular":
                return export_sales_report(results["data"])
            elif report_req_info["export"]["type"] == "gstr1":
                return export_gstr1_report(report_req_info["filters"]["invoiceDate"]["start"], report_req_info["filters"]["invoiceDate"]["end"])
                
        return results
    elif report_req_info["reportType"] == "purchase":
        results = get_purchase_report(report_req_info["filters"], 
                            report_req_info["sorters"], 
                            report_req_info["pageRequest"], 
                            report_req_info["maxItemsPerPage"])
        if (report_req_info["export"]["required"]):
            return export_purchase_report(results["data"])
        return results

def export_sales_report(invoices):
    file_base_dir = "./tempdata/sales_report/"
    filename = str(datetime.datetime.now())+"sales_report.xlsx"
    wb = openpyxl.Workbook() 
    sheet = wb.active
    sheet.title = "Sales Report"
    sheet.freeze_panes = 'A2'

    column_headers = [
        "Invoice No.", 
        "Invoice Date", 
        "Customer Name", 
        "GSTIN", 
        "Item Description", 
        "Item Code", 
        "HSN", 
        "Qty", 
        "Rate per Item", 
        "Taxable Val", 
        "CGST Rate", 
        "CGST Amt", 
        "SGST Rate", 
        "SGST Amt", 
        "IGST Rate", 
        "IGST Amt", 
        "Total",
        "size"]

    item_keys = [
        "itemDesc",
        "itemCode",
        "HSN",
        "quantity",
        "ratePerItem",
        "taxableValue",
        "CGST",
        "CGSTAmount",
        "SGST",
        "SGSTAmount",
        "IGST",
        "IGSTAmount",
        "value",
    ]

    for i, column_header in enumerate(column_headers):
        sheet.cell(row=1, column=i+1).value = column_header

    #find a better way to do the following, instead of using base index, use sheet.max_row ?
    base_index = 0
    for invoice in invoices:
        gst_tables = compute_gst_tables(invoice.productItems, invoice.serviceItems)
        if (invoice.customerDetails.GSTIN == "" or invoice.customerDetails.GSTIN.startswith("09")):
            tax_table = gst_tables["GST_table"]
        else:
            tax_table = gst_tables["IGST_table"]
        products = tax_table["products"]
        services = tax_table["services"]

        total_items = len(products) + len(services)
        for i, product in enumerate(products):
            row_index = i+2+base_index
            sheet.cell(row = row_index, column = 1).value = invoice.invoiceNumber
            sheet.cell(row = row_index, column = 2).value = invoice.invoiceDate.strftime("%d/%m/%Y")
            sheet.cell(row = row_index, column = 3).value = invoice.customerDetails.name
            sheet.cell(row = row_index, column = 4).value = invoice.customerDetails.GSTIN

            for j, item_key in enumerate(item_keys):
                sheet.cell(row = row_index, column = 5+j).value = product[item_key]
            
            sheet.cell(row = row_index, column = 5+len(item_keys)).value = compute_size(product["itemDesc"])
            
        for i, service in enumerate(services):
            row_index = i+2+base_index+len(products)
            sheet.cell(row = row_index, column = 1).value = invoice.invoiceNumber
            sheet.cell(row = row_index, column = 2).value = invoice.invoiceDate.strftime("%d/%m/%Y")
            sheet.cell(row = row_index, column = 3).value = invoice.customerDetails.name
            sheet.cell(row = row_index, column = 4).value = invoice.customerDetails.GSTIN

            for j, item_key in enumerate(item_keys):
                sheet.cell(row = row_index, column = 5+j).value = service[item_key]

        base_index += total_items

    total_row_index = sheet.max_row
    
    # total row label
    sheet.cell(row=total_row_index+1, column=7).value = "TOTAL"
    # total quantity
    total_quantity = '= SUM(H2:H'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=8).value = total_quantity
    # total taxable val
    total_taxable_val = '= SUM(J2:J'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=10).value = total_taxable_val
    # total CGST
    total_CGST = '= SUM(L2:L'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=12).value = total_CGST
    # total SGST
    total_SGST = '= SUM(N2:N'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=14).value = total_SGST
    # total IGST
    total_IGST = '= SUM(P2:P'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=16).value = total_IGST
    # total 
    total = '= SUM(Q2:Q'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=17).value = total

    for i in range(2, sheet.max_row+1):
        sheet.cell(row=i, column=11).number_format = FORMAT_PERCENTAGE
        sheet.cell(row=i, column=13).number_format = FORMAT_PERCENTAGE
        sheet.cell(row=i, column=15).number_format = FORMAT_PERCENTAGE

    #this adjusts the font and column width of the whole sheet
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), (len(str(cell.value)))*1.2))    
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value
    ####

    # Apply styles to header and footer row
    for col in range(1, sheet.max_column + 1):
        cell_header = sheet.cell(1, col)
        cell_header.fill = PatternFill("solid", fgColor="b4f05c")
        cell_header.border = Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        cell_header.font = Font(name="Calibri", bold=True)
        cell_footer = sheet.cell(sheet.max_row, col)
        cell_footer.fill = PatternFill("solid", fgColor="b4f05c")
        cell_footer.border = Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        cell_footer.font = Font(name="Calibri", bold=True)

    wb.save(file_base_dir+filename)
    return filename

def get_gst_state_codes():
	file_path = "./gstStateCodes.json"
	f = open(file_path, "r")
	gst_state_codes = json.loads(f.read())
	f.close()
	return gst_state_codes

# helper function for export_gstr1_report
def round_to_two(sheet, value):
    cell = WriteOnlyCell(sheet, value=round(value, 2))
    cell.number_format = FORMAT_NUMBER_00
    return cell

def export_gstr1_report(start_date, end_date):
    file_base_dir = "./tempdata/sales_report/"
    filename = str(datetime.datetime.now())+"gstr1.xlsx"
    wb = openpyxl.Workbook(write_only=True)
    gst_state_codes = get_gst_state_codes()
    gstr1_report = prepare_gstr1_report(start_date, end_date)
    # ----------------------------------------------b2b sheet----------------------------------------------
    b2b_sheet = wb.create_sheet("b2b,sez,de")
    b2b_summary = gstr1_report["b2bData"]["b2bSummary"]
    b2b_items = gstr1_report["b2bData"]["b2bItems"]

    excel_b2b_summary = {
            "No. of Recipients" : b2b_summary["countUniqB2bCustomer"],
            "No. of Invoices": b2b_summary["countUniqB2bInvoices"],
            "Total Invoice Value": round_to_two(b2b_sheet, b2b_summary["sumInvoiceTotal"]),
            "Total Taxable Value": round_to_two(b2b_sheet, b2b_summary["sumTaxableValue"]),
            "Total Cess": round_to_two(b2b_sheet, 0.00)
        }
    excel_b2b_items = []
    for b2b_item in b2b_items:
        excel_b2b_items.append({
            "GSTIN/UIN of Recipient": b2b_item["customerGSTIN"],
            "Receiver Name": b2b_item["customerName"],
            "Invoice Number": b2b_item["invoiceNumber"],
            "Invoice date": b2b_item["invoiceDate"].strftime("%d-%b-%Y"), 
            "Invoice Value": round_to_two(b2b_sheet, b2b_item["invoiceTotal"]), 
            "Place Of Supply": f"09-{gst_state_codes['09']}", 
            "Reverse Charge": "N", 
            "Applicable % of Tax Rate": "", 
            "Invoice Type": "Regular B2B", 
            "E-Commerce GSTIN": "", 
            "Rate": round_to_two(b2b_sheet, b2b_item["itemRate"]), 
            "Taxable Value": round_to_two(b2b_sheet, b2b_item["taxableValue"]),
            "Cess Amount": round_to_two(b2b_sheet, 0.00)
        })

    title_row = ["Summary For B2B, SEZ, DE (4A, 4B, 6B, 6C)"] + ["" for i in range(11)] + ["HELP"]
    summary_headers = ( ["No. of Recipients", "", "No. of Invoices", "", "Total Invoice Value"] 
                        + ["" for i in range(6)] 
                        + ["Total Taxable Value", "Total Cess"])
    summary_data = ([excel_b2b_summary["No. of Recipients"], "", excel_b2b_summary["No. of Invoices"], "", 
                     excel_b2b_summary["Total Invoice Value"] ]
                    +["" for i in range(6)]
                    +[excel_b2b_summary["Total Taxable Value"], excel_b2b_summary["Total Cess"]])
    column_headers = list(excel_b2b_items[0])

    b2b_sheet.append(title_row)
    b2b_sheet.append(summary_headers)
    b2b_sheet.append(summary_data)
    b2b_sheet.append(column_headers)
    for excel_b2b_item in excel_b2b_items:
        b2b_sheet.append(list(excel_b2b_item.values()))

    # ----------------------------------------------b2cs sheet----------------------------------------------
    b2c_sheet = wb.create_sheet("b2cs")
    b2c_summary = gstr1_report["b2cData"]["b2cSummary"]
    b2c_items = gstr1_report["b2cData"]["b2cItems"]
    
    excel_b2c_summary = {
            "Total Taxable  Value" : round_to_two(b2c_sheet, b2c_summary["sumTaxableValue"]),
            "Total Cess": round_to_two(b2c_sheet, 0.00)
        }
    excel_b2c_items = []
    for b2c_item in b2c_items:
        excel_b2c_items.append({
            "Type": "OE",
            "Place Of Supply": f"09-{gst_state_codes['09']}", 
            "Applicable % of Tax Rate": "", 
            "Rate": round_to_two(b2c_sheet, b2c_item["rate"]), 
            "Taxable Value": round_to_two(b2c_sheet, b2c_item["taxableValue"]),
            "Cess Amount": round_to_two(b2c_sheet, 0.00),
            "E-Commerce GSTIN": ""
        })

    title_row = ["Summary For B2CS(7)"] + ["" for i in range(5)] + ["HELP"]

    summary_headers = ["" for i in range(4)] + ["Total Taxable  Value", "Total Cess", ""]
    summary_data = ["" for i in range(4)] + [excel_b2c_summary["Total Taxable  Value"], excel_b2c_summary["Total Cess"], ""]
    column_headers = list(excel_b2c_items[0])

    b2c_sheet.append(title_row)
    b2c_sheet.append(summary_headers)
    b2c_sheet.append(summary_data)
    b2c_sheet.append(column_headers)
    for excel_b2c_item in excel_b2c_items:
        b2c_sheet.append(list(excel_b2c_item.values()))

# ----------------------------------------------hsn sheet----------------------------------------------
    hsn_sheet = wb.create_sheet("hsn")
    hsn_summary = gstr1_report["hsnData"]["hsnSummary"]
    hsn_items = gstr1_report["hsnData"]["hsnItems"]
    
    excel_hsn_summary = {
            "No. of HSN" : hsn_summary["countUniqHSN"],
            "Total Value": "",
            "Total Taxable Value": round_to_two(hsn_sheet, hsn_summary["sumTaxableValue"]),
            "Total Integrated Tax": round_to_two(hsn_sheet, hsn_summary["sumIGSTAmount"]),
            "Total Central Tax": round_to_two(hsn_sheet, hsn_summary["sumCGSTAmount"]),
            "Total State/UT Tax": round_to_two(hsn_sheet, hsn_summary["sumSGSTAmount"]),
            "Total Cess": round_to_two(hsn_sheet, 0.00)
        }
    excel_hsn_items = []
    for item in hsn_items:
        excel_hsn_items.append({
            "HSN": item["hsn"],
            "Description": "", 
            "UQC": "NOS-NUMBERS", 
            "Total Quantity": round_to_two(hsn_sheet, item["quantity"]), 
            "Total Value": "",
            "Rate": round_to_two(hsn_sheet, item["rate"]),
            "Taxable Value": round_to_two(hsn_sheet, item["taxableValue"]),
            "Integrated Tax Amount": round_to_two(hsn_sheet, item["IGSTAmount"]),
            "Central Tax Amount" : round_to_two(hsn_sheet, item["CGSTAmount"]),
            "State/UT Tax Amount": round_to_two(hsn_sheet, item["SGSTAmount"]),
            "Cess Amount": round_to_two(hsn_sheet, 0.00)
        })

    title_row = ["Summary For HSN(12)"] + ["" for i in range(9)] + ["HELP"]
    summary_headers = [
        "No. of HSN", 
        "", 
        "", 
        "", 
        "Total Value", 
        "", 
        "Total Taxable Value", 
        "Total Integrated Tax",
        "Total Central Tax",
        "Total State/UT Tax",
        "Total Cess"
        ]
    summary_data = [
        excel_hsn_summary["No. of HSN"], 
        "", 
        "", 
        "", 
        excel_hsn_summary["Total Value"], 
        "", 
        excel_hsn_summary["Total Taxable Value"], 
        excel_hsn_summary["Total Integrated Tax"],
        excel_hsn_summary["Total Central Tax"],
        excel_hsn_summary["Total State/UT Tax"],
        excel_hsn_summary["Total Cess"]
        ]
    column_headers = list(excel_hsn_items[0])

    hsn_sheet.append(title_row)
    hsn_sheet.append(summary_headers)
    hsn_sheet.append(summary_data)
    hsn_sheet.append(column_headers)
    for excel_hsn_item in excel_hsn_items:
        hsn_sheet.append(list(excel_hsn_item.values()))

# ----------------------------------------------docs sheet----------------------------------------------
    docs_sheet = wb.create_sheet("docs")
    docs_summary = gstr1_report["docsData"]["docsSummary"]
    docs_items = gstr1_report["docsData"]["docsItems"]
    
    excel_docs_summary = {
            "Total Number" : docs_summary["totalDocs"],
            "Total Cancelled": docs_summary["totalcancelledDocs"],
        }
    excel_docs_items = []
    for item in docs_items:
        excel_docs_items.append({
            "Nature of Document": "Invoices for outward supply",
            "Sr. No. From": item["minInvoiceNumber"], 
            "Sr. No. To": item["maxInvoiceNumber"], 
            "Total Number": item["paidInvoicesCount"], 
            "Cancelled": item["cancelledInvoicesCount"],
        })

    title_row = ["Summary of documents issued during the tax period (13)"] + ["" for i in range(3)] + ["HELP"]
    summary_headers = [
        "", 
        "", 
        "", 
        "Total Number", 
        "Total Cancelled", 
        ]
    summary_data = [
        "", 
        "", 
        "", 
        excel_docs_summary["Total Number"], 
        excel_docs_summary["Total Cancelled"], 
        ]
    column_headers = list(excel_docs_items[0])

    docs_sheet.append(title_row)
    docs_sheet.append(summary_headers)
    docs_sheet.append(summary_data)
    docs_sheet.append(column_headers)
    for excel_docs_item in excel_docs_items:
        docs_sheet.append(list(excel_docs_item.values()))

    wb.save(file_base_dir+filename)
    return filename

def export_purchase_report(invoices):
    file_base_dir = "./tempdata/sales_report/"
    filename = str(datetime.datetime.now())+"purchase_report.xlsx"
    wb = openpyxl.Workbook() 
    sheet = wb.active

    sheet.freeze_panes = 'A2'

    column_headers = ["Invoice No.", "Invoice Date", "Claim Invoice", "Item Description", "Item Code", "HSN", "Qty", "Taxable Val", "Tax", "Total", "Size"]
    
    for i, column_header in enumerate(column_headers):
        sheet.cell(row=1, column=i+1).value = column_header

    #find a better way to do the following, instead of using base index, use sheet.max_row ?
    base_index = 0
    for invoice in invoices:
        total_items = len(invoice.items)
        for i, product in enumerate(invoice.items):
            row_index = i+2+base_index
            sheet.cell(row = row_index, column = 1).value = invoice.invoiceNumber
            sheet.cell(row = row_index, column = 2).value = invoice.invoiceDate.strftime("%d/%m/%Y")
            sheet.cell(row = row_index, column = 3).value = "No" if invoice.claimInvoice ==0 else "Yes"
            sheet.cell(row = row_index, column = 4).value = product.itemDesc
            sheet.cell(row = row_index, column = 5).value = product.itemCode
            sheet.cell(row = row_index, column = 6).value = int(product.HSN)
            sheet.cell(row = row_index, column = 7).value = product.quantity
            sheet.cell(row = row_index, column = 8).value = product.taxableValue
            sheet.cell(row = row_index, column = 9).value = product.tax
            sheet.cell(row = row_index, column = 10).value = product.itemTotal
            sheet.cell(row = row_index, column = 11).value = compute_size(product.itemDesc)
        base_index += total_items

    total_row_index = sheet.max_row
    
    # total row label
    sheet.cell(row=total_row_index+1, column=2).value = "TOTAL"
    # total quantity
    total_quantity = '= SUM(G2:G'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=7).value = total_quantity
    # total taxable val
    total_taxable_val = '= SUM(H2:H'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=8).value = total_taxable_val
    # total CGST
    total_tax = '= SUM(I2:I'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=9).value = total_tax
    # total 
    total = '= SUM(J2:J'+str(total_row_index)+')'
    sheet.cell(row=total_row_index+1, column=10).value = total

    for i in range(2, sheet.max_row+1):
        sheet.cell(row=i, column=6).number_format = FORMAT_NUMBER

    #this adjusts the font and column width of the whole sheet
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), (len(str(cell.value))+1)*1.2))    
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value
    ####

    # Apply styles to header and footer row
    for col in range(1, sheet.max_column + 1):
        cell_header = sheet.cell(1, col)
        cell_header.fill = PatternFill("solid", fgColor="C5C5C5")
        cell_header.border = Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        cell_header.font = Font(name="Calibri", bold=True)
        cell_footer = sheet.cell(sheet.max_row, col)
        cell_footer.fill = PatternFill("solid", fgColor="C5C5C5")
        cell_footer.border = Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        cell_footer.font = Font(name="Calibri", bold=True)

    wb.save(file_base_dir+filename)
    return filename

def stock_report():
    category_required_list = [
    "passenger_car_tyre", 
    "passenger_car_tube", 
    "2_wheeler_tyre",
    "2_wheeler_tube",
    "3_wheeler_tyre",
    "3_wheeler_tube",
    "scv_tyre",
    "scv_tube",
    "tubeless_valve",
    "loose_tube/flaps_tube"
    ]
    products = Product.objects(Q(category__in=category_required_list))
    file_base_dir = "./tempdata/sales_report/"
    filename = str(datetime.datetime.now())+"stock_report.xlsx"
    wb = openpyxl.Workbook() 
    sheet = wb.active

    sheet.freeze_panes = 'A2'
    column_headers = ["Item Description", "Item Code", "HSN", "GST", "category", "Size", "Cost Price", "Stock"]
    product_keys = [
        "itemDesc",
        "itemCode",
        "HSN",
        "GST",
        "category",
        "size",
        "costPrice",
        "stock"
    ]

    for i, column_header in enumerate(column_headers):
        sheet.cell(row=1, column=i+1).value = column_header

    for i, product in enumerate(products):
        for j, product_key in enumerate(product_keys):
            sheet.cell(row = i+2, column = j+1).value = product[product_key]

    #this adjusts the font and column width of the whole sheet
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), (len(str(cell.value))+0)*1.2))    
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value
    ####

    # Apply styles to header and footer row
    for col in range(1, sheet.max_column + 1):
        cell_header = sheet.cell(1, col)
        cell_header.fill = PatternFill("solid", fgColor="fcfa8e")
        cell_header.border = Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        cell_header.font = Font(name="Calibri", bold=True)

    wb.save(file_base_dir+filename)
    return filename

def reset_stock():
    for product in Product.objects():
        product.update(stock = 0)

    for invoice in Purchase.objects:
        for product in invoice.items:
            product_found = Product.objects(itemCode=product.itemCode).first()
            if product_found is None:
                print(f'Item from Purchase Invoice No. {invoice.invoiceNumber} not found in Product table:  {product.itemDesc}, {product.itemCode}, ')
                return False
            new_stock = product_found.stock + product.quantity
            Product.objects(itemCode=product.itemCode).first().update(stock=new_stock) 
    
    for invoice in Sale.objects:
        for product in invoice.productItems:
            product_found = Product.objects(itemCode=product.itemCode).first()
            if product_found is None:
                print(f'Item from Sale Invoice No. {invoice.invoiceNumber} not found in Product table:  {product.itemDesc}, {product.itemCode}, ')
                return False
            new_stock = product_found.stock - product.quantity
            Product.objects(itemCode=product.itemCode).first().update(stock=new_stock)
    return True    